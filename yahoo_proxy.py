#!/usr/bin/env python3
"""
Yahoo Fantasy OAuth Proxy for NHL Roster Tracker
-------------------------------------------------
Run once:  python3 yahoo_proxy.py
Then click "Sync Yahoo" in the roster tracker.

Requires: pip install requests flask flask-cors
"""

import json, os, re, time, threading, webbrowser, ssl, subprocess, sys, traceback, unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs, urlencode
import requests
from flask import Flask, jsonify, Response, request
from flask_cors import CORS

# ── Config ────────────────────────────────────────────────────────────────────
# Load credentials from a .env file if present, then fall back to environment variables.
_env_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env_file):
    with open(_env_file) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _, _v = _line.partition("=")
                os.environ.setdefault(_k.strip(), _v.strip())

CLIENT_ID = os.environ.get("YAHOO_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("YAHOO_CLIENT_SECRET", "")

if not CLIENT_ID or not CLIENT_SECRET:
    sys.exit(
        "❌  YAHOO_CLIENT_ID and YAHOO_CLIENT_SECRET must be set.\n"
        "    Create a .env file next to yahoo_proxy.py with:\n\n"
        "        YAHOO_CLIENT_ID=your_client_id\n"
        "        YAHOO_CLIENT_SECRET=your_client_secret\n\n"
        "    Or export them as environment variables before running."
    )
LEAGUE_KEY = os.environ.get("YAHOO_LEAGUE_KEY", "465.l.97882")
REDIRECT_URI = "https://localhost:8100/callback"
TOKEN_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".yahoo_tokens.json")
CERT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".localhost_cert.pem")
KEY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".localhost_key.pem")
PROXY_PORT = 8099
CALLBACK_PORT = 8100

YAHOO_AUTH_URL = "https://api.login.yahoo.com/oauth2/request_auth"
YAHOO_TOKEN_URL = "https://api.login.yahoo.com/oauth2/get_token"
YAHOO_API_BASE = "https://fantasysports.yahooapis.com/fantasy/v2"


# ── SSL cert generation ───────────────────────────────────────────────────────
def ensure_cert():
    if os.path.exists(CERT_FILE) and os.path.exists(KEY_FILE):
        return
    print("🔐  Generating self-signed SSL cert for localhost…")
    subprocess.run([
        "openssl", "req", "-x509", "-newkey", "rsa:2048",
        "-keyout", KEY_FILE, "-out", CERT_FILE,
        "-days", "825", "-nodes",
        "-subj", "/CN=localhost",
        "-addext", "subjectAltName=IP:127.0.0.1,DNS:localhost"
    ], check=True, capture_output=True)
    print("✅  Cert generated.")


# ── Token storage ─────────────────────────────────────────────────────────────
def load_tokens():
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE) as f:
            return json.load(f)
    return {}


def save_tokens(tokens):
    with open(TOKEN_FILE, "w") as f:
        json.dump(tokens, f, indent=2)


def token_expired(tokens):
    return time.time() > tokens.get("expires_at", 0) - 60


def refresh_access_token(tokens):
    print("🔄  Refreshing Yahoo access token…")
    resp = requests.post(YAHOO_TOKEN_URL, data={
        "grant_type": "refresh_token",
        "refresh_token": tokens["refresh_token"],
        "redirect_uri": REDIRECT_URI,
    }, auth=(CLIENT_ID, CLIENT_SECRET))
    resp.raise_for_status()
    data = resp.json()
    tokens["access_token"] = data["access_token"]
    tokens["refresh_token"] = data.get("refresh_token", tokens["refresh_token"])
    tokens["expires_at"] = time.time() + data.get("expires_in", 3600)
    save_tokens(tokens)
    print("✅  Token refreshed.")
    return tokens


_token_lock = threading.Lock()


def get_valid_token():
    with _token_lock:
        tokens = load_tokens()
        if not tokens:
            return None
        if token_expired(tokens):
            tokens = refresh_access_token(tokens)
        return tokens["access_token"]


# ── OAuth callback server ─────────────────────────────────────────────────────
auth_code_holder = {}


class CallbackHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        qs = parse_qs(urlparse(self.path).query)
        code = qs.get("code", [None])[0]
        if code:
            auth_code_holder["code"] = code
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.end_headers()
            self.wfile.write(b"""
                <html><body style="font-family:sans-serif;padding:2rem;text-align:center;background:#f9f9f9">
                <div style="max-width:400px;margin:4rem auto;background:#fff;padding:2rem;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.08)">
                <div style="font-size:48px">&#10003;</div>
                <h2 style="color:#185FA5;margin:.5rem 0">Authorized!</h2>
                <p style="color:#666">You can close this tab and return to the roster tracker.<br>Click <strong>Sync Yahoo</strong> to load your league.</p>
                </div></body></html>
            """)
        else:
            self.send_response(400)
            self.end_headers()
            self.wfile.write(b"Missing code.")

    def log_message(self, *args):
        pass


def exchange_code(code):
    resp = requests.post(YAHOO_TOKEN_URL, data={
        "grant_type": "authorization_code",
        "code": code,
        "redirect_uri": REDIRECT_URI,
    }, auth=(CLIENT_ID, CLIENT_SECRET))
    resp.raise_for_status()
    data = resp.json()
    tokens = {
        "access_token": data["access_token"],
        "refresh_token": data["refresh_token"],
        "expires_at": time.time() + data.get("expires_in", 3600),
    }
    save_tokens(tokens)
    return tokens


def do_oauth_flow():
    params = {
        "client_id": CLIENT_ID,
        "redirect_uri": REDIRECT_URI,
        "response_type": "code",
        "scope": "fspt-r",
    }
    auth_url = YAHOO_AUTH_URL + "?" + urlencode(params)
    print(f"\n🌐  Opening Yahoo authorization in your browser…")
    print(f"    If it doesn't open automatically, visit:\n    {auth_url}\n")
    webbrowser.open(auth_url)

    # HTTPS callback server
    server = HTTPServer(("localhost", CALLBACK_PORT), CallbackHandler)
    ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
    ctx.load_cert_chain(certfile=CERT_FILE, keyfile=KEY_FILE)
    server.socket = ctx.wrap_socket(server.socket, server_side=True)
    server.timeout = 120

    print("⏳  Waiting for Yahoo callback (up to 2 minutes)…")
    print(f"    ⚠️  Your browser may warn about the self-signed cert at the callback URL.")
    print(f"    If it does: click Advanced → Proceed to localhost (unsafe)\n")

    while "code" not in auth_code_holder:
        server.handle_request()
    server.server_close()

    code = auth_code_holder["code"]
    print("🔑  Got authorization code — exchanging for tokens…")
    exchange_code(code)
    print("✅  Tokens saved! You're all set.\n")


# ── Yahoo API helper ──────────────────────────────────────────────────────────
def yahoo_get(path, token):
    sep = "&" if "?" in path else "?"
    url = YAHOO_API_BASE + path + sep + "format=json"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    return resp.json()


# ── Flask proxy ───────────────────────────────────────────────────────────────
app = Flask(__name__)
CORS(app, origins=["*"])


@app.route("/status")
def status():
    tokens = load_tokens()
    if not tokens:
        return jsonify({"authenticated": False})
    return jsonify({"authenticated": True, "expired": token_expired(tokens)})


@app.route("/auth")
def auth():
    threading.Thread(target=do_oauth_flow, daemon=True).start()
    return jsonify({"ok": True, "message": "OAuth flow started — check your browser."})


@app.route("/league")
def league():
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        data = yahoo_get(f"/league/{LEAGUE_KEY}", token)
        return jsonify({"ok": True, "league": data["fantasy_content"]["league"][0]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/teamlogos")
def teamlogos():
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        data = yahoo_get(f"/league/{LEAGUE_KEY}/teams", token)
        teams_raw = data["fantasy_content"]["league"][1]["teams"]
        logos = {}
        for idx in (k for k in teams_raw.keys() if k != "count"):
            team_meta = teams_raw[idx]["team"][0]
            name = next((m["name"] for m in team_meta if isinstance(m, dict) and "name" in m), None)
            logo_entry = next((m for m in team_meta if isinstance(m, dict) and "team_logos" in m), None)
            if name and logo_entry:
                logos[name] = logo_entry["team_logos"][0]["team_logo"]["url"]
        return jsonify({"ok": True, "logos": logos})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _fetch_rosters(league_key, token):
    """Fetch all rostered players for a given Yahoo league key. Returns a list of player dicts."""
    data = yahoo_get(f"/league/{league_key}/teams/roster/players", token)
    teams_raw = data["fantasy_content"]["league"][1]["teams"]
    result = []
    for idx in (k for k in teams_raw.keys() if k != "count"):
        team = teams_raw[idx]["team"]
        team_name = next((m["name"] for m in team[0] if isinstance(m, dict) and "name" in m), f"Team {idx}")
        roster_data = team[1]["roster"]
        if isinstance(roster_data, list):
            players_raw = next(r["players"] for r in roster_data if isinstance(r, dict) and "players" in r)
        else:
            players_raw = next(v["players"] for k, v in roster_data.items() if isinstance(v, dict) and "players" in v)
        for pidx in (k for k in players_raw.keys() if k != "count"):
            p    = players_raw[pidx]["player"]
            meta = p[0]
            full_name = next((m.get("full_name") or m.get("name", {}).get("full", "")
                              for m in meta if isinstance(m, dict) and ("full_name" in m or "name" in m)), "")
            nhl_team  = next((m.get("editorial_team_abbr", "—").upper()
                              for m in meta if isinstance(m, dict) and "editorial_team_abbr" in m), "—")
            positions = next((m.get("eligible_positions", [])
                              for m in meta if isinstance(m, dict) and "eligible_positions" in m), [])
            pos_list  = [ep.get("position", "") for ep in positions if isinstance(ep, dict)]
            pos       = next((p for p in pos_list if p in ("C","LW","RW","D","G","F")), "") or ""
            if full_name:
                result.append({"name": full_name, "nhlTeam": nhl_team, "pos": pos, "fantasyTeam": team_name})
    return result


LEAGUE_KEY_2425 = "453.l.52799"
LEAGUE_KEY_2324 = "427.l.1827"
LEAGUE_KEY_2223 = "419.l.52810"


@app.route("/teams-2324")
def teams_2324():
    """Try candidate NHL game keys against league 1827 to find the 23-24 key."""
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    # NHL game keys near 2023-24 (453=24-25, so 23-24 is somewhere below)
    candidates = [LEAGUE_KEY_2324] + [f"{g}.l.1827" for g in [427, 428, 429, 430, 431, 432, 433, 434, 435, 436, 437, 438, 439, 440, 441, 442]]
    results = {}
    for key in candidates:
        try:
            data = yahoo_get(f"/league/{key}/teams", token)
            teams_raw = data["fantasy_content"]["league"][1]["teams"]
            names = sorted([
                next((m["name"] for m in teams_raw[k]["team"][0] if isinstance(m, dict) and "name" in m), f"Team {k}")
                for k in teams_raw if k != "count"
            ])
            results[key] = {"ok": True, "teams": names}
            print(f"  ✓ {key}: {names}")
        except Exception as e:
            results[key] = {"error": str(e)[:80]}
            print(f"  ✗ {key}: {e}")
    return jsonify(results)


@app.route("/rosters")
def rosters():
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        return jsonify({"ok": True, "players": _fetch_rosters(LEAGUE_KEY, token)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/rosters-2324")
def rosters_2324():
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        players = _fetch_rosters(LEAGUE_KEY_2324, token)
        print(f"  2023-24 rosters: {len(players)} players fetched")
        return jsonify({"ok": True, "players": players})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/rosters-2223")
def rosters_2223():
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        players = _fetch_rosters(LEAGUE_KEY_2223, token)
        print(f"  2022-23 rosters: {len(players)} players fetched")
        return jsonify({"ok": True, "players": players})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/rosters-2425")
def rosters_2425():
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        players = _fetch_rosters(LEAGUE_KEY_2425, token)
        print(f"  2024-25 rosters: {len(players)} players fetched")
        return jsonify({"ok": True, "players": players})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _get_league_team_names(league_key, token):
    """Return {team_key: team_name} for a league."""
    data = yahoo_get(f"/league/{league_key}/teams", token)
    teams_raw = data["fantasy_content"]["league"][1]["teams"]
    result = {}
    for k in [x for x in teams_raw if x != "count"]:
        row = teams_raw[k]["team"][0]
        tkey  = next((m["team_key"] for m in row if isinstance(m, dict) and "team_key" in m), None)
        tname = next((m["name"]     for m in row if isinstance(m, dict) and "name"     in m), f"Team {k}")
        if tkey:
            result[tkey] = tname
    return result


@app.route("/yahoo-draftresults")
def yahoo_draftresults():
    """Return all pre-season draft picks for a league, with player names resolved.

    ?league_key=  defaults to current LEAGUE_KEY.
    Calls /draftresults then batch-fetches player metadata.
    """
    league_key = request.args.get("league_key", LEAGUE_KEY)
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        team_names = _get_league_team_names(league_key, token)

        data = yahoo_get(f"/league/{league_key}/draftresults", token)
        dr   = data["fantasy_content"]["league"][1]["draftresults"]
        count = int(dr.get("count", 0))

        raw_picks   = []
        player_keys = []
        for i in range(count):
            pick = dr.get(str(i), {}).get("draft_pick", {})
            if pick:
                raw_picks.append(pick)
                if pick.get("player_key"):
                    player_keys.append(pick["player_key"])

        # Resolve player names in batches of 25
        player_info = {}
        for i in range(0, len(player_keys), 25):
            batch = ",".join(player_keys[i:i + 25])
            try:
                pd = yahoo_get(f"/players;player_keys={batch};out=metadata", token)
                pr = pd["fantasy_content"]["players"]
                for j in range(int(pr.get("count", 0))):
                    p = pr.get(str(j), {}).get("player", [])
                    if not p:
                        continue
                    meta = p[0]
                    pkey  = next((m["player_key"] for m in meta if isinstance(m, dict) and "player_key" in m), None)
                    pname = next((m.get("full_name") or m.get("name", {}).get("full", "")
                                  for m in meta if isinstance(m, dict) and ("full_name" in m or "name" in m)), "")
                    ppos  = next((m.get("display_position", "")
                                  for m in meta if isinstance(m, dict) and "display_position" in m), "")
                    if pkey:
                        player_info[pkey] = {"name": pname, "pos": ppos}
            except Exception as e:
                print(f"  draftresults player batch error: {e}")
            time.sleep(0.1)

        picks = [
            {
                "round":  int(p.get("round", 0)),
                "pick":   int(p.get("pick",  0)),
                "team":   team_names.get(p.get("team_key", ""), p.get("team_key", "")),
                "player": player_info.get(p.get("player_key", ""), {}).get("name", p.get("player_key", "")),
                "pos":    player_info.get(p.get("player_key", ""), {}).get("pos", "?"),
            }
            for p in raw_picks
        ]
        picks.sort(key=lambda x: x["pick"])
        print(f"  draftresults: {len(picks)} picks for {league_key}")
        return jsonify({"ok": True, "picks": picks, "count": len(picks)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/yahoo-drops")
def yahoo_drops():
    """Return drop transactions for a league.

    ?league_key=  defaults to current LEAGUE_KEY.
    ?before=YYYY-MM-DD  only returns drops before this date (use to isolate
                        pre-draft drops from in-season drops).
    Paginates through all transactions of type=drop.
    """
    import datetime as _dt
    league_key = request.args.get("league_key", LEAGUE_KEY)
    before_str = request.args.get("before")
    before_ts  = None
    if before_str:
        try:
            before_ts = _dt.datetime.strptime(before_str, "%Y-%m-%d").timestamp()
        except ValueError:
            pass
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        team_names = _get_league_team_names(league_key, token)
        all_drops  = []
        start, batch = 0, 100

        while True:
            data = yahoo_get(
                f"/league/{league_key}/transactions;type=drop;count={batch};start={start}", token
            )
            tx_raw   = data["fantasy_content"]["league"][1].get("transactions", {})
            returned = int(tx_raw.get("count", 0))

            for i in [k for k in tx_raw if k != "count"]:
                tx = tx_raw[i].get("transaction")
                if not tx or len(tx) < 2:
                    continue
                ts = int(tx[0].get("timestamp", 0))
                players_raw = tx[1].get("players", {})
                for j in [k for k in players_raw if k != "count"]:
                    pp = players_raw[j].get("player")
                    if not pp or len(pp) < 2:
                        continue
                    meta = pp[0]
                    pname = next((m.get("full_name") or m.get("name", {}).get("full", "")
                                  for m in meta if isinstance(m, dict) and ("full_name" in m or "name" in m)), "")
                    tx_data = pp[1].get("transaction_data", {})
                    src_key = ""
                    if isinstance(tx_data, list) and tx_data:
                        src_key = tx_data[0].get("source_team_key", "")
                    elif isinstance(tx_data, dict):
                        src_key = tx_data.get("source_team_key", "")
                    if pname:
                        all_drops.append({
                            "timestamp": ts,
                            "team":   team_names.get(src_key, src_key),
                            "player": pname,
                        })

            if returned < batch:
                break
            start += batch
            time.sleep(0.1)

        all_drops.sort(key=lambda x: x["timestamp"])
        if before_ts:
            all_drops = [d for d in all_drops if d["timestamp"] < before_ts]
        print(f"  drops: {len(all_drops)} for {league_key}{f' before {before_str}' if before_str else ''}")
        return jsonify({"ok": True, "drops": all_drops, "count": len(all_drops)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/yahoo-xactions-debug")
def yahoo_xactions_debug():
    """Return raw Yahoo transaction response (first 5) for structure inspection.
    Tries several query variants to find which returns data."""
    league_key = request.args.get("league_key", LEAGUE_KEY)
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    results = {}
    variants = [
        ("add,drop typed",    f"/league/{league_key}/transactions;type=add,drop;count=5;start=0"),
        ("no type filter",    f"/league/{league_key}/transactions;count=5;start=0"),
        ("add only",          f"/league/{league_key}/transactions;type=add;count=5;start=0"),
        ("drop only",         f"/league/{league_key}/transactions;type=drop;count=5;start=0"),
    ]
    for label, url in variants:
        try:
            data = yahoo_get(url, token)
            league1 = data["fantasy_content"]["league"][1]
            tx = league1.get("transactions", "MISSING") if isinstance(league1, dict) else f"league1 is {type(league1).__name__}"
            results[label] = {
                "url": url,
                "transactions_type": type(tx).__name__,
                "transactions_len_or_keys": len(tx) if isinstance(tx, (list, dict)) else str(tx),
                "sample": tx if not isinstance(tx, (list, dict)) else (list(tx.items())[:2] if isinstance(tx, dict) else tx[:2]),
            }
        except Exception as e:
            results[label] = {"error": str(e)}
    return jsonify(results)


@app.route("/yahoo-xactions")
def yahoo_xactions():
    """Return all add/drop transactions for a league, paginated.

    ?league_key=  defaults to current LEAGUE_KEY.
    Returns [{timestamp, date, team, player, pos, type: 'add'|'drop'}] sorted by date.

    Structure (confirmed via API + Apps Script reference):
    - Call with no type filter — type=add,drop returns []
    - Only process outer type == "add/drop" (combined waiver moves); pure drops are
      preseason-only and show up with different structure, handled separately below
    - transaction_data is list [{...}] for adds, dict {...} for drops
    - Team name is in destination_team_name / source_team_name
    """
    import datetime as _dt
    league_key = request.args.get("league_key", LEAGUE_KEY)
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        all_tx = []
        start, batch = 0, 100

        while True:
            data = yahoo_get(
                f"/league/{league_key}/transactions;count={batch};start={start}", token
            )
            tx_raw = data["fantasy_content"]["league"][1].get("transactions", {})

            if isinstance(tx_raw, list):
                break  # [] means no transactions accessible for this league
            returned = int(tx_raw.get("count", 0))
            if returned == 0:
                break

            for k in [k for k in tx_raw if k != "count"]:
                tx = tx_raw[k].get("transaction") if isinstance(tx_raw[k], dict) else None
                if not tx or len(tx) < 2:
                    continue

                meta = tx[0] if isinstance(tx[0], dict) else {}
                tx_type = meta.get("type", "")
                ts = int(meta.get("timestamp", 0))
                date_str = _dt.datetime.fromtimestamp(ts).strftime("%Y-%m-%d") if ts else ""

                players_raw = tx[1].get("players", {}) if isinstance(tx[1], dict) else {}
                for j in sorted(k for k in players_raw if k != "count"):
                    pp = players_raw[j].get("player") if isinstance(players_raw[j], dict) else None
                    if not pp or len(pp) < 2:
                        continue

                    # pp[0] = list of meta dicts: player_key, player_id, name{full,first,...}, display_position, ...
                    pmeta = pp[0] if isinstance(pp[0], list) else []
                    pname = pos = ""
                    for m in pmeta:
                        if not isinstance(m, dict):
                            continue
                        if "name" in m:
                            nm = m["name"]
                            pname = nm.get("full", "") if isinstance(nm, dict) else str(nm)
                        if "display_position" in m:
                            pos = m["display_position"]

                    # pp[1] = {"transaction_data": list-or-dict}
                    pp1 = pp[1] if isinstance(pp[1], dict) else {}
                    raw_td = pp1.get("transaction_data", {})
                    td = raw_td[0] if isinstance(raw_td, list) and raw_td else raw_td
                    if not isinstance(td, dict):
                        continue

                    ptype = td.get("type", "")
                    if ptype == "add":
                        team = td.get("destination_team_name", td.get("destination_team_key", ""))
                    elif ptype == "drop":
                        team = td.get("source_team_name", td.get("source_team_key", ""))
                    else:
                        continue

                    if pname and team:
                        all_tx.append({
                            "timestamp": ts,
                            "date":      date_str,
                            "team":      team,
                            "player":    pname,
                            "pos":       pos,
                            "type":      ptype,
                        })

            if returned < batch:
                break
            start += batch
            time.sleep(0.1)

        all_tx.sort(key=lambda x: x["timestamp"])
        print(f"  xactions: {len(all_tx)} for {league_key}")
        return jsonify({"ok": True, "transactions": all_tx, "count": len(all_tx)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/transactions")
def transactions():
    token = get_valid_token()
    if not token:
        return jsonify({"error": "Not authenticated."}), 401
    try:
        data = yahoo_get(f"/league/{LEAGUE_KEY}/transactions;type=add,drop,trade;count=25", token)
        tx_raw = data["fantasy_content"]["league"][1].get("transactions", {})
        result = []
        tx_indices = [k for k in tx_raw.keys() if k != "count"]
        for i in tx_indices:
            tx = tx_raw[i].get("transaction")
            if not tx: continue
            meta = tx[0]
            ts = int(meta.get("timestamp", 0))
            date_str = time.strftime("%-m/%-d", time.localtime(ts)) if ts else ""
            players_raw = tx[1].get("players", {})
            players_list = []
            p_indices = [k for k in players_raw.keys() if k != "count"]
            for j in p_indices:
                pp = players_raw[j].get("player")
                if not pp: continue
                p_meta = pp[0]
                name = next((m.get("full_name") or m.get("name", {}).get("full", "")
                             for m in p_meta if isinstance(m, dict) and ("full_name" in m or "name" in m)), "")
                tx_data = pp[1].get("transaction_data", {})
                ptype = tx_data[0].get("type", meta.get("type", "")) if isinstance(tx_data,
                                                                                   list) and tx_data else meta.get(
                    "type", "")
                if name:
                    players_list.append({"name": name, "type": ptype})
            if players_list:
                result.append({"date": date_str, "type": meta.get("type", ""), "players": players_list, "ts": ts})
        result.sort(key=lambda x: x["ts"], reverse=True)
        return jsonify({"ok": True, "transactions": result})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/")
def index():
    html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "roster_tracker.html")
    if os.path.exists(html_path):
        with open(html_path) as f:
            return Response(f.read(), mimetype="text/html")
    return Response("<h2>roster_tracker.html not found — place it in the same folder as yahoo_proxy.py</h2>",
                    mimetype="text/html")


@app.route("/muller.webp")
def muller_image():
    from flask import send_file as _sf
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "muller.webp")
    return _sf(path, mimetype="image/webp")


@app.route("/data_2324.js")
def data_2324_js():
    from flask import send_file as _sf
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_2324.js")
    return _sf(path, mimetype="application/javascript")


@app.route("/data_2425.js")
def data_2425_js():
    from flask import send_file as _sf
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_2425.js")
    return _sf(path, mimetype="application/javascript")


@app.route("/data_2526.js")
def data_2526_js():
    from flask import send_file as _sf
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_2526.js")
    return _sf(path, mimetype="application/javascript")


@app.route("/data_2627.js")
def data_2627_js():
    from flask import send_file as _sf
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_2627.js")
    return _sf(path, mimetype="application/javascript")


# ── Team lists ────────────────────────────────────────────────────────────────
# Used for PuckPedia cap-hit scraping (ARI kept for legacy depth-chart pages)
NHL_TEAMS = [
    "ANA", "ARI", "BOS", "BUF", "CAR", "CBJ", "CGY", "CHI", "COL", "DAL",
    "DET", "EDM", "FLA", "LAK", "MIN", "MTL", "NJD", "NSH", "NYI", "NYR",
    "OTT", "PHI", "PIT", "SEA", "SJS", "STL", "TBL", "TOR", "UTA", "VAN",
    "VGK", "WSH", "WPG"
]
# ARI no longer exists in the NHL API (relocated → UTA)
NHL_API_TEAMS = [t for t in NHL_TEAMS if t != "ARI"]


# NHL positionCode → app abbreviation
_POS_MAP = {"C": "C", "L": "LW", "R": "RW", "D": "D", "G": "G"}


def normalize_name(name: str) -> str:
    """Fold accented chars to ASCII and strip hyphens/periods/apostrophes for key matching.
    Must stay in sync with JS normName: s.normalize('NFD').replace(/[̀-ͯ]/g,'').replace(/[-.']/g,'')"""
    nfd = unicodedata.normalize("NFD", name)
    ascii_only = nfd.encode("ascii", "ignore").decode("ascii")
    return ascii_only.replace("-", "").replace(".", "").replace("'", "")


def _pos_group(pos):
    """Map any position abbreviation to F / D / G."""
    if pos == "D":
        return "D"
    if pos == "G":
        return "G"
    return "F"


# ── NHL roster status ─────────────────────────────────────────────────────────
@app.route("/nhlrosters")
def nhl_rosters():
    """Returns a dict keyed by lowercased full name for all players on active NHL rosters."""
    active = {}
    errors = []

    for team in NHL_API_TEAMS:
        try:
            url = f"https://api-web.nhle.com/v1/roster/{team}/current"
            resp = requests.get(url, timeout=10)
            resp.raise_for_status()
            for group in ("forwards", "defensemen", "goalies"):
                for p in resp.json().get(group, []):
                    first = p.get("firstName", {}).get("default", "")
                    last = p.get("lastName", {}).get("default", "")
                    full = f"{first} {last}".strip()
                    pos = _POS_MAP.get(p.get("positionCode", ""), "C")
                    if full:
                        key = f"{normalize_name(full).lower()}_{_pos_group(pos)}"
                        if key in active:
                            print(f"  ⚠ duplicate: {full!r} ({team}) already seen on {active[key]['nhlTeam']}, skipping")
                            continue
                        active[key] = {"name": full, "nhlTeam": team, "pos": pos,
                                       "birthDate": p.get("birthDate", "")}
                        _add_name_aliases(active, key, full, _pos_group(pos))
            print(f"  ✓ {team}")
        except Exception as e:
            errors.append(f"{team}: {str(e)}")
            print(f"  ✗ {team}: {e}")
        time.sleep(0.3)

    return jsonify({"ok": True, "active": active, "count": len(active), "errors": errors})


# ── 2025-26 season stats ──────────────────────────────────────────────────────

_NHL_STATS_BASE = "https://api.nhle.com/stats/rest/en"


@app.route("/stats")
def season_stats():
    """
    Returns 2025-26 regular-season fantasy points for all NHL players.
    Scoring: Goal=1, Assist=1, Goalie Win=2, Goalie Shutout=3 (bonus on top of win).
    Uses the NHL REST stats API to fetch all skaters/goalies in two calls.
    """
    stats = {}
    errors = []
    cayenne = "seasonId=20252026%20and%20gameTypeId=2"

    try:
        start, page_size, fetched = 0, 100, []
        while True:
            url = f"{_NHL_STATS_BASE}/skater/summary?limit={page_size}&start={start}&cayenneExp={cayenne}"
            page = requests.get(url, timeout=30).json()
            batch = page.get("data", [])
            fetched.extend(batch)
            if len(fetched) >= page.get("total", 0) or not batch:
                break
            start += page_size
        for s in fetched:
            full = s.get("skaterFullName", "").strip()
            if not full:
                continue
            pos = _POS_MAP.get(s.get("positionCode", ""), "C")
            key = f"{normalize_name(full).lower()}_{_pos_group(pos)}"
            fpts = s.get("goals", 0) + s.get("assists", 0)
            stats[key] = {
                "name": full, "fpts": fpts,
                "goals": s.get("goals", 0), "assists": s.get("assists", 0),
                "gp": s.get("gamesPlayed", 0), "ppp": s.get("ppPoints", 0),
            }
            _add_name_aliases(stats, key, full, _pos_group(pos))
        print(f"  2025-26 NHL skaters: {len(fetched)} fetched")
    except Exception as e:
        errors.append(f"skaters: {e}")
        print(f"  ✗ skaters: {e}")

    try:
        url = f"{_NHL_STATS_BASE}/goalie/summary?limit=200&cayenneExp={cayenne}"
        data = requests.get(url, timeout=30).json()
        for g in data.get("data", []):
            full = g.get("goalieFullName", "").strip()
            if not full:
                continue
            key = f"{normalize_name(full).lower()}_G"
            wins = g.get("wins", 0)
            sos  = g.get("shutouts", 0)
            stats[key] = {
                "name": full, "fpts": wins * 2 + sos * 3,
                "wins": wins, "shutouts": sos,
                "gp": g.get("gamesPlayed", 0), "ppp": 0,
            }
            _add_name_aliases(stats, key, full, "G")
        print(f"  2025-26 NHL goalies: {len(data.get('data', []))} fetched")
    except Exception as e:
        errors.append(f"goalies: {e}")
        print(f"  ✗ goalies: {e}")

    print(f"Stats done: {len(stats)} players, {len(errors)} errors")
    return jsonify({"ok": True, "players": stats, "count": len(stats), "errors": errors})


@app.route("/stats-2627")
def season_stats_2627():
    """Returns 2026-27 regular-season fantasy points for all NHL players."""
    stats = {}
    errors = []
    cayenne = "seasonId=20262027%20and%20gameTypeId=2"

    try:
        start, page_size, fetched = 0, 100, []
        while True:
            url = f"{_NHL_STATS_BASE}/skater/summary?limit={page_size}&start={start}&cayenneExp={cayenne}"
            page = requests.get(url, timeout=30).json()
            batch = page.get("data", [])
            fetched.extend(batch)
            if len(fetched) >= page.get("total", 0) or not batch:
                break
            start += page_size
        for s in fetched:
            full = s.get("skaterFullName", "").strip()
            if not full:
                continue
            pos = _POS_MAP.get(s.get("positionCode", ""), "C")
            key = f"{normalize_name(full).lower()}_{_pos_group(pos)}"
            fpts = s.get("goals", 0) + s.get("assists", 0)
            stats[key] = {
                "name": full, "fpts": fpts,
                "goals": s.get("goals", 0), "assists": s.get("assists", 0),
                "gp": s.get("gamesPlayed", 0), "ppp": s.get("ppPoints", 0),
            }
            _add_name_aliases(stats, key, full, _pos_group(pos))
    except Exception as e:
        errors.append(f"skaters: {e}")

    try:
        url = f"{_NHL_STATS_BASE}/goalie/summary?limit=200&cayenneExp={cayenne}"
        data = requests.get(url, timeout=30).json()
        for g in data.get("data", []):
            full = g.get("goalieFullName", "").strip()
            if not full:
                continue
            key = f"{normalize_name(full).lower()}_G"
            wins = g.get("wins", 0)
            sos  = g.get("shutouts", 0)
            stats[key] = {
                "name": full, "fpts": wins * 2 + sos * 3,
                "wins": wins, "shutouts": sos,
                "gp": g.get("gamesPlayed", 0), "ppp": 0,
            }
            _add_name_aliases(stats, key, full, "G")
    except Exception as e:
        errors.append(f"goalies: {e}")

    return jsonify({"ok": True, "players": stats, "count": len(stats), "errors": errors})


@app.route("/stats-2324")
def season_stats_2324():
    """Returns 2023-24 regular-season fantasy points for all NHL players."""
    stats = {}
    errors = []
    cayenne = "seasonId=20232024%20and%20gameTypeId=2"

    try:
        start, page_size, fetched = 0, 100, []
        while True:
            url = f"{_NHL_STATS_BASE}/skater/summary?limit={page_size}&start={start}&cayenneExp={cayenne}"
            page = requests.get(url, timeout=30).json()
            batch = page.get("data", [])
            fetched.extend(batch)
            if len(fetched) >= page.get("total", 0) or not batch:
                break
            start += page_size
        for s in fetched:
            full = s.get("skaterFullName", "").strip()
            if not full:
                continue
            pos = _POS_MAP.get(s.get("positionCode", ""), "C")
            key = f"{normalize_name(full).lower()}_{_pos_group(pos)}"
            fpts = s.get("goals", 0) + s.get("assists", 0)
            stats[key] = {
                "name": full, "fpts": fpts,
                "goals": s.get("goals", 0), "assists": s.get("assists", 0),
                "gp": s.get("gamesPlayed", 0), "ppp": s.get("ppPoints", 0),
            }
            _add_name_aliases(stats, key, full, _pos_group(pos))
        print(f"  2023-24 NHL skaters: {len(fetched)} fetched")
    except Exception as e:
        errors.append(f"skaters: {e}")
        print(f"  ✗ skaters: {e}")

    try:
        url = f"{_NHL_STATS_BASE}/goalie/summary?limit=200&cayenneExp={cayenne}"
        data = requests.get(url, timeout=30).json()
        for g in data.get("data", []):
            full = g.get("goalieFullName", "").strip()
            if not full:
                continue
            key = f"{normalize_name(full).lower()}_G"
            wins = g.get("wins", 0)
            sos  = g.get("shutouts", 0)
            stats[key] = {
                "name": full, "fpts": wins * 2 + sos * 3,
                "wins": wins, "shutouts": sos,
                "gp": g.get("gamesPlayed", 0), "ppp": 0,
            }
            _add_name_aliases(stats, key, full, "G")
        print(f"  2023-24 NHL goalies: {len(data.get('data', []))} fetched")
    except Exception as e:
        errors.append(f"goalies: {e}")
        print(f"  ✗ goalies: {e}")

    print(f"Stats 2023-24 done: {len(stats)} players, {len(errors)} errors")
    return jsonify({"ok": True, "players": stats, "count": len(stats), "errors": errors})


@app.route("/stats-2425")
def season_stats_2425():
    """Returns 2024-25 regular-season fantasy points for all NHL players."""
    stats = {}
    errors = []
    cayenne = "seasonId=20242025%20and%20gameTypeId=2"

    try:
        start, page_size, fetched = 0, 100, []
        while True:
            url = f"{_NHL_STATS_BASE}/skater/summary?limit={page_size}&start={start}&cayenneExp={cayenne}"
            page = requests.get(url, timeout=30).json()
            batch = page.get("data", [])
            fetched.extend(batch)
            if len(fetched) >= page.get("total", 0) or not batch:
                break
            start += page_size
        for s in fetched:
            full = s.get("skaterFullName", "").strip()
            if not full:
                continue
            pos = _POS_MAP.get(s.get("positionCode", ""), "C")
            key = f"{normalize_name(full).lower()}_{_pos_group(pos)}"
            fpts = s.get("goals", 0) + s.get("assists", 0)
            stats[key] = {
                "name": full, "fpts": fpts,
                "goals": s.get("goals", 0), "assists": s.get("assists", 0),
                "gp": s.get("gamesPlayed", 0), "ppp": s.get("ppPoints", 0),
            }
            _add_name_aliases(stats, key, full, _pos_group(pos))
        print(f"  2024-25 NHL skaters: {len(fetched)} fetched")
    except Exception as e:
        errors.append(f"skaters: {e}")
        print(f"  ✗ skaters: {e}")

    try:
        url = f"{_NHL_STATS_BASE}/goalie/summary?limit=200&cayenneExp={cayenne}"
        data = requests.get(url, timeout=30).json()
        for g in data.get("data", []):
            full = g.get("goalieFullName", "").strip()
            if not full:
                continue
            key = f"{normalize_name(full).lower()}_G"
            wins = g.get("wins", 0)
            sos  = g.get("shutouts", 0)
            stats[key] = {
                "name": full, "fpts": wins * 2 + sos * 3,
                "wins": wins, "shutouts": sos,
                "gp": g.get("gamesPlayed", 0), "ppp": 0,
            }
            _add_name_aliases(stats, key, full, "G")
        print(f"  2024-25 NHL goalies: {len(data.get('data', []))} fetched")
    except Exception as e:
        errors.append(f"goalies: {e}")
        print(f"  ✗ goalies: {e}")

    print(f"Stats 2024-25 done: {len(stats)} players, {len(errors)} errors")
    return jsonify({"ok": True, "players": stats, "count": len(stats), "errors": errors})


@app.route("/signings-2425")
def signings_2425():
    """Return each player's active cap hit as of the 24-25 draft date (Oct 5, 2024).

    Queries PuckPedia for contracts signed on or before 2024-10-05 whose
    expiry covers at least the 2024-25 season. First occurrence per player
    (sign_date DESC) is their 24-25 cap hit.
    """
    import urllib.parse

    try:
        scraper = _make_puckpedia_scraper()
        all_players = {}
        errors = []

        SEASON_CUTOFF = "2024-2025"
        DATE_FROM = "2017-01-01"
        DATE_TO   = "2024-10-05"

        q_base = {
            "pageSize": 100,
            "sortBy": "sign_date",
            "sortDirection": "DESC",
            "sign_date": [DATE_FROM, DATE_TO],
        }

        q_base["curPage"] = 1
        resp = scraper.get(
            "https://puckpedia.com/data/api_signings?q=" + urllib.parse.quote(json.dumps(q_base)),
            timeout=20,
        )
        resp.raise_for_status()
        first = resp.json()["data"]
        total_count = first["meta"]["count"]
        total_pages = (total_count + 99) // 100
        print(f"Signings 24-25: {total_count} records, {total_pages} pages")

        def process_page(records):
            for p in records:
                exp = p.get("exp", "")
                if exp < SEASON_CUTOFF:
                    continue
                term = int(p.get("len") or 0)
                exp_year = int(exp.split("-")[0]) if exp else 0
                start_yr = exp_year - term + 1 if term else 0
                if start_yr > 2024:
                    continue
                pos = p.get("pos", "C")
                pg = _POS_GROUP_MAP.get(pos, "F")
                full = f"{p.get('p_fn','')} {p.get('p_ln','')}".strip()
                if not full:
                    continue
                first_lower = p.get("p_fn", "").lower()
                nick = _FORMAL_TO_NICK.get(first_lower)
                key = f"{normalize_name(full).lower()}_{pg}"
                nick_key = f"{normalize_name(nick + ' ' + p.get('p_ln','')).lower()}_{pg}" if nick else None
                cap = round(float(p["cap_hit"]) / 1_000_000, 4)
                if key not in all_players:
                    entry = {"name": full, "pos": pos, "cap": cap, "nhl_id": p.get("p_nhl_id", "")}
                    all_players[key] = entry
                    if nick_key and nick_key not in all_players:
                        all_players[nick_key] = entry

        process_page(first["p"])

        for page in range(2, total_pages + 1):
            try:
                q_base["curPage"] = page
                r = scraper.get(
                    "https://puckpedia.com/data/api_signings?q=" + urllib.parse.quote(json.dumps(q_base)),
                    timeout=20,
                )
                r.raise_for_status()
                process_page(r.json()["data"]["p"])
                print(f"  page {page}/{total_pages} ({len(all_players)} active so far)")
                time.sleep(0.15)
            except Exception as e:
                errors.append(f"page {page}: {str(e)}")
                print(f"  ✗ page {page}: {e}")

        print(f"Signings 24-25 done: {len(all_players)} players, {len(errors)} errors")
        return jsonify({"ok": True, "players": all_players, "count": len(all_players), "errors": errors})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── ESPN Fantasy Hockey projections ──────────────────────────────────────────

_ESPN_HDRS = {
    "Accept":  "application/json",
    "Referer": "https://fantasy.espn.com/hockey/players/projections?leagueFormatId=2",
}
_ESPN_POS  = {1: "F", 2: "F", 3: "F", 4: "D", 5: "G"}   # defaultPositionId → group


def _espn_api_url(season_year: int) -> str:
    return (f"https://lm-api-reads.fantasy.espn.com/apis/v3/games/fhl"
            f"/seasons/{season_year}/segments/0/leaguedefaults/2?view=kona_player_info")


def _espn_fetch_group(slot_ids: list, season_year: int = 2026) -> list:
    """Page through ESPN API for the given slot IDs and season, return all raw player dicts."""
    api_url = _espn_api_url(season_year)
    results = []
    offset  = 0
    total   = None
    while total is None or offset < total:
        ff = json.dumps({
            "players": {
                "limit": 50,
                "offset": offset,
                "sortPercOwned": {"sortPriority": 1, "sortAsc": False},
                "filterSlotIds": {"value": slot_ids},
            }
        })
        r = requests.get(api_url, headers={**_ESPN_HDRS, "X-Fantasy-Filter": ff}, timeout=15)
        r.raise_for_status()
        if total is None:
            total = int(r.headers.get("X-Fantasy-Filter-Player-Count", 0))
        results.extend(r.json()["players"])
        offset += 50
        if offset < total:
            time.sleep(0.15)
    return results


def _espn_fetch_projections(season_year: int):
    """Fetch HGHL fantasy-point projections from ESPN for the given season year."""
    all_players = {}
    errors      = []
    groups = [
        ([0, 1, 2, 3, 4, 6], False, "skaters"),
        ([5],                 True,  "goalies"),
    ]
    for slot_ids, is_goalie, label in groups:
        try:
            raw = _espn_fetch_group(slot_ids, season_year)
            for entry in raw:
                pl   = entry["player"]
                name = pl.get("fullName", "")
                if not name:
                    continue
                pg = _ESPN_POS.get(pl.get("defaultPositionId"), "F")
                proj = next(
                    (s for s in pl.get("stats", [])
                     if s.get("statSourceId") == 1
                     and s.get("statSplitTypeId") == 0
                     and s.get("seasonId") == season_year
                     and s.get("stats")),
                    None,
                )
                if not proj:
                    continue
                st = proj["stats"]
                if is_goalie:
                    hghl_pts = round(st.get("1", 0) * 2 + st.get("7", 0) * 3)
                    ppp = 0
                else:
                    hghl_pts = round(st.get("16") or (st.get("13", 0) + st.get("14", 0)))
                    ppp = round(st.get("19", 0) or st.get("17", 0) + st.get("18", 0))
                gp = round(st.get("30", 0))  # ESPN's projected-games-played stat id (confirmed for both skaters/goalies)
                if not all_players:
                    print(f"  ESPN stat keys (first player '{name}'): {sorted(st.keys())}")
                key = f"{normalize_name(name).lower()}_{pg}"
                all_players[key] = {"name": name, "pg": pg, "hghl_pts": hghl_pts, "ppp": ppp, "gp": gp}
                _add_name_aliases(all_players, key, name, pg)
            print(f"  ESPN {season_year} {label}: {len(raw)} fetched")
        except Exception as e:
            errors.append(f"{label}: {str(e)}")
            print(f"  ✗ ESPN {season_year} {label}: {e}")
    return all_players, errors


@app.route("/espn-projections")
def espn_projections():
    """Fetch ESPN 2026-27 HGHL fantasy-point projections."""
    players, errors = _espn_fetch_projections(2026)
    return jsonify({"ok": True, "players": players, "count": len(players), "errors": errors})


@app.route("/espn-projections-2526")
def espn_projections_2526():
    """Fetch ESPN 2025-26 HGHL fantasy-point projections (historical)."""
    players, errors = _espn_fetch_projections(2025)
    return jsonify({"ok": True, "players": players, "count": len(players), "errors": errors})


# ── DFO / 5v5hockey projections scraper ──────────────────────────────────────

_DFO_URL = "https://5v5hockey.com/projections-embedded/"
_DFO_HDR = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}


def _parse_dfo_func(func_name: str, html: str) -> list:
    """Extract player records from an inline JS function in the 5v5hockey page."""
    idx = html.find(f"function {func_name}")
    if idx == -1:
        return []
    start = html.find("const data = [", idx)
    if start == -1:
        return []
    end = html.find("\n      return data", start)
    section = html[start: end if end > 0 else start + 500_000]

    players = []
    for m in re.finditer(r"player:\s*\{'logo':[^,]+,\s*'name':\s*'([^']+)'", section):
        name = m.group(1)
        block_start = section.rfind("{", 0, m.start())
        block_end   = section.find("\n          },", m.end())
        block = section[block_start: block_end if block_end > 0 else block_start + 2000]

        def _fld(key):
            fm = re.search(rf"\b{key}:\s*([\d.]+)", block)
            return float(fm.group(1)) if fm else 0.0

        pos_m = re.search(r'Pos:\s*"([^"]+)"', block)
        players.append({
            "name": name,
            "pos":  pos_m.group(1) if pos_m else "",
            "G": _fld("G"), "A": _fld("A"), "PPP": _fld("PPP"),
            "W": _fld("W"), "SO": _fld("SO"),
        })
    return players


@app.route("/dfo-projections")
def dfo_projections():
    """Scrape 5v5hockey/DFO projections (Points league). G+A for skaters, W×2+SO×3 for goalies."""
    try:
        r = requests.get(_DFO_URL, headers=_DFO_HDR, timeout=30)
        r.raise_for_status()
        html = r.text
    except Exception as e:
        return jsonify({"error": f"Could not fetch DFO page: {e}"}), 500

    skaters = _parse_dfo_func("rowsSkatersPts", html)
    goalies = _parse_dfo_func("rowGoaliesPts",  html)

    all_players = {}
    for p in skaters:
        hghl_pts = round(p["G"] + p["A"])
        if hghl_pts <= 0:
            continue
        pg  = "D" if p["pos"] == "D" else "F"
        key = f"{normalize_name(p['name']).lower()}_{pg}"
        all_players[key] = {"name": p["name"], "pg": pg, "hghl_pts": hghl_pts, "ppp": round(p["PPP"])}
        _add_name_aliases(all_players, key, p["name"], pg)
    for p in goalies:
        hghl_pts = round(p["W"] * 2 + p["SO"] * 3)
        if hghl_pts <= 0:
            continue
        key = f"{normalize_name(p['name']).lower()}_G"
        all_players[key] = {"name": p["name"], "pg": "G", "hghl_pts": hghl_pts, "ppp": 0}
        _add_name_aliases(all_players, key, p["name"], "G")

    print(f"  DFO: {len(skaters)} skaters + {len(goalies)} goalies → {len(all_players)} total")
    return jsonify({"ok": True, "players": all_players, "count": len(all_players)})


# ── CBS Sports projections scraper ───────────────────────────────────────────

_CBS_HDR = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"}
_CBS_POS = {"C": "F", "LW": "F", "RW": "F", "D": "D", "G": "G"}


def _cbs_parse_rows(html: str) -> list:
    """Extract (name, [numeric stat values]) tuples from a CBS projections table page."""
    start = html.find("<table")
    end = html.find("</table>", start)
    if start == -1 or end == -1:
        return []
    table = html[start:end]
    rows = []
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", table, re.S):
        name_m = re.search(r'class="fn">([^<]+)</span>', row)
        if not name_m:
            continue
        team_m = re.search(r'class="team">([^<]+)</span>', row)
        vals = [float(v) for v in re.findall(r'<td class="numeric">([\d.-]+)</td>', row)]
        rows.append((name_m.group(1).strip(), team_m.group(1).strip() if team_m else "", vals))
    return rows


def _cbs_fetch_position(pos: str) -> list:
    """Page through CBS Sports projections for a given position (C/LW/RW/D/G)."""
    rows = []
    page = 1
    max_page = 1
    while True:
        url = f"https://www.cbssports.com/fantasy/hockey/projections/{pos}/" + (f"{page}/" if page > 1 else "")
        r = requests.get(url, headers=_CBS_HDR, timeout=20)
        r.raise_for_status()
        rows.extend(_cbs_parse_rows(r.text))
        if page == 1:
            pages = [int(m) for m in re.findall(rf"/fantasy/hockey/projections/{pos}/(\d+)/", r.text)]
            max_page = max(pages) if pages else 1
        if page >= max_page:
            break
        page += 1
    return rows


def _cbs_fetch_projections():
    """Fetch CBS Sports 2026-27 fantasy hockey projections for all positions.
    Skater rows: GP, G, A, PTS, +/-, PPG, S, SHG, SPct, PIM, FPTS (11 values) → G+A for HGHL pts.
    Goalie rows: GGP, W, L, SO, GAA, GA, S, SOGA, SPct, Min, FPTS (11 values) → W*2+SO*3 for HGHL pts.
    """
    all_players = {}
    errors = []
    for pos, pg in _CBS_POS.items():
        try:
            rows = _cbs_fetch_position(pos)
        except Exception as e:
            errors.append(f"CBS {pos}: {e}")
            continue
        for name, team, vals in rows:
            if len(vals) < 11:
                continue
            if pg == "G":
                hghl_pts = round(vals[1] * 2 + vals[3] * 3)
            else:
                hghl_pts = round(vals[1] + vals[2])
            if hghl_pts <= 0:
                continue
            key = f"{normalize_name(name).lower()}_{pg}"
            all_players[key] = {"name": name, "pg": pg, "team": team, "hghl_pts": hghl_pts}
            _add_name_aliases(all_players, key, name, pg)
        print(f"  CBS {pos}: {len(rows)} players")
    return all_players, errors


@app.route("/cbs-projections")
def cbs_projections():
    """Fetch CBS Sports 2026-27 fantasy hockey projections (G+A for skaters, W*2+SO*3 for goalies)."""
    players, errors = _cbs_fetch_projections()
    print(f"  CBS: {len(players)} total")
    return jsonify({"ok": True, "players": players, "count": len(players), "errors": errors})


_NHL_PROJ_URLS = {
    "2526": "https://www.nhl.com/news/nhl-fantasy-hockey-projections-forward-defenseman-points-2025-26",
    "2627": "https://www.nhl.com/news/nhl-fantasy-hockey-projections-forward-defenseman-points-2026-27",
}

def _parse_nhl_proj_page(html: str) -> dict:
    """Parse NHL.com fantasy projections page text into {normName_pos: pts}."""
    # Unescape literal \n sequences in the embedded JSON blob so they become real newlines
    text = html.replace('\\n', '\n')
    # Skip optional leading HTML tag (e.g. <p>) before the name
    pattern = re.compile(
        r'^(?:<[^>]+>)*([A-Z][^,\n<]{1,30}),\s*([FD]),\s*[A-Z]{2,4}[^:\n]*:\s*(\d+)',
        re.MULTILINE
    )
    players = {}
    for m in pattern.finditer(text):
        name, pos, pts = m.group(1).strip(), m.group(2).strip(), int(m.group(3))
        key = f"{normalize_name(name).lower()}_{pos}"
        players[key] = pts
    return players


@app.route("/nhl-projections-2526")
def nhl_projections_2526():
    """Fetch NHL.com 2025-26 forward/defenseman point projections."""
    try:
        resp = requests.get(_NHL_PROJ_URLS["2526"], timeout=15,
                            headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        players = _parse_nhl_proj_page(resp.content.decode("utf-8"))
        return jsonify({"ok": True, "players": players, "count": len(players)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/nhl-projections-2627")
def nhl_projections_2627():
    """Fetch NHL.com 2026-27 forward/defenseman point projections."""
    try:
        resp = requests.get(_NHL_PROJ_URLS["2627"], timeout=15,
                            headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        players = _parse_nhl_proj_page(resp.content.decode("utf-8"))
        return jsonify({"ok": True, "players": players, "count": len(players)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/dfo-projections-2526", methods=["POST"])
def dfo_projections_2526():
    """Parse a DFO/5v5hockey CSV export for 2025-26 projections.
    Expected columns: Player, (dup), Team, VAR, ADP, Pos, Site Pos, GP, G, A, PTS, ..., Points
    The last 'Points' column is the pre-calculated HGHL total (G+A for skaters, W*2+SO*3 for goalies).
    Pos column: C/LW/RW → F, D → D, G → G.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".csv"):
        return jsonify({"error": "Please upload a .csv file"}), 400
    try:
        import csv, io
        text = f.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
    except Exception as e:
        return jsonify({"error": f"Could not read CSV: {e}"}), 400

    if not rows:
        return jsonify({"error": "Empty CSV"}), 400

    header = [h.strip() for h in rows[0]]
    try:
        col_name = 0
        col_pos  = header.index("Pos")
        col_pts  = len(header) - 1  # last column = "Points" (HGHL total)
    except (ValueError, IndexError) as e:
        return jsonify({"error": f"Could not find expected columns: {e}"}), 400
    col_ppp = header.index("PPP") if "PPP" in header else None

    def safe_float(val):
        try: return float(val) if val and val.strip() else 0.0
        except: return 0.0

    all_players = {}
    for row in rows[1:]:
        if not row or not row[col_name].strip():
            continue
        name = row[col_name].strip()
        pos_raw = row[col_pos].strip().upper() if len(row) > col_pos else ""
        if not pos_raw:
            continue
        pg = "G" if pos_raw == "G" else "D" if pos_raw == "D" else "F"
        hghl_pts = round(safe_float(row[col_pts]) if len(row) > col_pts else 0)
        if hghl_pts <= 0:
            continue
        ppp = round(safe_float(row[col_ppp]) if col_ppp is not None and len(row) > col_ppp else 0)
        key = f"{normalize_name(name).lower()}_{pg}"
        all_players[key] = {"name": name, "pg": pg, "hghl_pts": hghl_pts, "ppp": ppp}
        _add_name_aliases(all_players, key, name, pg)

    print(f"  DFO 25-26 CSV: {len(all_players)} players parsed")
    return jsonify({"ok": True, "players": all_players, "count": len(all_players)})


@app.route("/nhl-projections-2425", methods=["POST"])
def nhl_projections_2425():
    """Parse the NHL 24-25 projections CSV.
    Columns: combined data, first, last, position, NHL team, points, ..., Proj Pts
    Name = first + last; position = D/F; hghl_pts = points (col 5 = G+A for skaters).
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".csv"):
        return jsonify({"error": "Please upload a .csv file"}), 400
    try:
        import csv, io
        text = f.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
    except Exception as e:
        return jsonify({"error": f"Could not read CSV: {e}"}), 400

    if not rows:
        return jsonify({"error": "Empty CSV"}), 400

    header = [h.strip().lower() for h in rows[0]]
    try:
        c_first = header.index("first")
        c_last  = header.index("last")
        c_pos   = header.index("position")
        c_pts   = header.index("points")
    except ValueError as e:
        return jsonify({"error": f"Could not find expected columns (first/last/position/points): {e}"}), 400

    def safe_float(v):
        try: return float(v) if v and str(v).strip() else 0.0
        except: return 0.0

    all_players = {}
    for row in rows[1:]:
        if len(row) <= max(c_first, c_last, c_pos, c_pts):
            continue
        first = row[c_first].strip()
        last  = row[c_last].strip()
        if not first or not last:
            continue
        name    = f"{first} {last}"
        pos_raw = row[c_pos].strip().upper()
        pg      = "G" if pos_raw == "G" else "D" if pos_raw == "D" else "F"
        hghl_pts = round(safe_float(row[c_pts]))
        if hghl_pts <= 0:
            continue
        key = f"{normalize_name(name).lower()}_{pg}"
        all_players[key] = {"name": name, "pg": pg, "hghl_pts": hghl_pts}

    print(f"  NHL 24-25 proj CSV: {len(all_players)} players parsed")
    return jsonify({"ok": True, "players": all_players, "count": len(all_players)})


@app.route("/athletic-projections-2526", methods=["POST"])
def athletic_projections_2526():
    """Same parser as /athletic-projections but tagged as 2025-26 historical data."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an .xlsx file"}), 400
    try:
        import openpyxl
        wb    = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sheet = wb["The List"] if "The List" in wb.sheetnames else wb.worksheets[0]
        rows  = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400

    header_idx = next((i for i, row in enumerate(rows)
                       if any(str(c).strip().upper() == "NAME" for c in row if c)), None)
    if header_idx is None:
        return jsonify({"error": "Could not find NAME column in header row"}), 400

    headers = [str(c).strip().upper() if c else "" for c in rows[header_idx]]

    def col(name):
        try: return headers.index(name)
        except ValueError: return None

    gp_all = [i for i, h in enumerate(headers) if h == "GP"]
    c_name, c_pos = col("NAME"), col("POS")
    c_g,  c_a     = col("G"),   col("A"); c_ppp = col("PPP")
    c_w,  c_so    = col("W"),   col("SO")
    c_gp_s = gp_all[0] if gp_all else None
    c_gp_g = gp_all[1] if len(gp_all) > 1 else None

    def safe(val):
        try: return float(val) if val is not None else 0.0
        except: return 0.0

    all_players = {}
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row): continue
        name = row[c_name] if c_name is not None else None
        if not name or str(name).strip().upper() == "NAME": continue
        name = str(name).strip()
        pos_raw = str(row[c_pos]).strip().upper() if c_pos is not None and row[c_pos] else ""
        if not pos_raw: continue
        pg = "G" if pos_raw == "G" else "D" if pos_raw == "D" else "F"
        if pg == "G":
            gp = safe(row[c_gp_g] if c_gp_g is not None else None)
            hghl_pts = round(safe(row[c_w]) * 2 + safe(row[c_so]) * 3)
            if gp <= 0 and hghl_pts <= 0: continue
            ppp = 0
        else:
            gp = safe(row[c_gp_s] if c_gp_s is not None else None)
            hghl_pts = round(safe(row[c_g]) + safe(row[c_a]))
            if gp <= 0 and hghl_pts <= 0: continue
            ppp = round(safe(row[c_ppp]) if c_ppp is not None else 0)
        if hghl_pts <= 0: continue
        key = f"{normalize_name(name).lower()}_{pg}"
        all_players[key] = {"name": name, "pg": pg, "hghl_pts": hghl_pts, "ppp": ppp, "gp": round(gp)}
        _add_name_aliases(all_players, key, name, pg)

    print(f"  Athletic 25-26: {len(all_players)} players parsed")
    return jsonify({"ok": True, "players": all_players, "count": len(all_players)})


# ── Athletic projections importer ────────────────────────────────────────────

@app.route("/athletic-projections", methods=["POST"])
def athletic_projections():
    """
    Parse The Athletic fantasy hockey Excel file ('The List' sheet).
    Expected columns (may include header rows above the data header):
      RK, NAME, KEEP?, POS, TEAM, AGE, SALARY, FP, /GP, VORP, /$, ADP, DIFF., ADJ,
      GP, TOI, G, A, PTS, SOG, PPG, PPP, SHG, SHP, BLK, HIT, +/-, PIM, GWG, FOW, FOL, FO%,
      GP, W, L, OTL, SO, SV, GA, SV%, GAA
    Skaters: HGHL pts = G + A  (projected season totals)
    Goalies: HGHL pts = W×2 + SO×3
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an .xlsx file (export from Numbers → File → Export To → Excel)"}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sheet = wb["The List"] if "The List" in wb.sheetnames else wb.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400

    # Find the header row (first row containing "NAME")
    header_idx = None
    for i, row in enumerate(rows):
        if any(str(c).strip().upper() == "NAME" for c in row if c is not None):
            header_idx = i
            break
    if header_idx is None:
        return jsonify({"error": "Could not find header row (expected a cell labelled NAME)"}), 400

    headers = [str(c).strip().upper() if c is not None else "" for c in rows[header_idx]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    # Two GP columns: first = skater, second = goalie
    gp_all = [i for i, h in enumerate(headers) if h == "GP"]
    c_name = col("NAME"); c_pos = col("POS")
    c_g = col("G");  c_a = col("A"); c_ppp = col("PPP")
    c_w = col("W");  c_so = col("SO")
    c_gp_s = gp_all[0] if len(gp_all) > 0 else None
    c_gp_g = gp_all[1] if len(gp_all) > 1 else None

    missing = [n for n, c in [("NAME", c_name), ("POS", c_pos), ("G", c_g),
                               ("A", c_a), ("W", c_w), ("SO", c_so)] if c is None]
    if missing:
        return jsonify({"error": f"Missing columns: {', '.join(missing)}"}), 400

    def safe(val):
        try:
            return float(val) if val is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    all_players = {}
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        name = row[c_name]
        if not name:
            continue
        name = str(name).strip()
        if not name or name.upper() == "NAME":
            continue

        pos_raw = str(row[c_pos]).strip().upper() if row[c_pos] else ""
        if not pos_raw:
            continue
        pg = "G" if pos_raw == "G" else "D" if pos_raw == "D" else "F"

        if pg == "G":
            gp = safe(row[c_gp_g] if c_gp_g is not None else None)
            w  = safe(row[c_w])
            so = safe(row[c_so])
            if gp <= 0 and w <= 0:
                continue
            hghl_pts = round(w * 2 + so * 3)
            ppp = 0
        else:
            gp = safe(row[c_gp_s] if c_gp_s is not None else None)
            g  = safe(row[c_g])
            a  = safe(row[c_a])
            if gp <= 0 and g <= 0 and a <= 0:
                continue
            hghl_pts = round(g + a)
            ppp = round(safe(row[c_ppp]) if c_ppp is not None else 0)

        if hghl_pts <= 0:
            continue

        key = f"{normalize_name(name).lower()}_{pg}"
        all_players[key] = {"name": name, "pg": pg, "hghl_pts": hghl_pts, "ppp": ppp, "gp": round(gp)}
        _add_name_aliases(all_players, key, name, pg)

    print(f"  Athletic: {len(all_players)} players parsed")
    return jsonify({"ok": True, "players": all_players, "count": len(all_players)})


# ── PuckPedia signings scraper ────────────────────────────────────────────────

def _make_puckpedia_scraper():
    """Return a curl_cffi session that impersonates Chrome's TLS fingerprint.

    cloudscraper fakes headers but uses Python's TLS stack, which Cloudflare
    fingerprints and blocks. curl_cffi replicates Chrome's actual TLS/HTTP2
    fingerprint at the libcurl level, which bypasses Cloudflare reliably.
    """
    from curl_cffi import requests as cffi_requests
    session = cffi_requests.Session(impersonate="chrome124")
    session.headers.update({
        "Referer":         "https://puckpedia.com/contracts",
        "Accept":          "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return session


_POS_GROUP_MAP = {"G": "G", "D": "D"}  # anything else → "F"

# PuckPedia uses legal/formal first names; NHL API and Yahoo use nicknames.
# Map formal → nickname so both keys are indexed in the signings lookup.
_FORMAL_TO_NICK = {
    "alexander": "alex",
    "alexandre": "alex",
    "bradley":   "brad",
    "cameron":   "cam",
    "christopher": "chris",
    "daniel":    "dan",
    "david":     "dave",
    "gabriel":   "gabe",
    "jacob":     "jake",
    "joshua":    "josh",
    "matthew":   "matt",
    "michael":   "mike",
    "mitchell":  "mitch",
    "nicholas":  "nick",
    "patrick":   "pat",
    "philip":    "phil",
    "richard":   "rick",
    "robert":    "rob",
    "samuel":    "sam",
    "thomas":    "tom",
    "timothy":   "tim",
    "william":   "will",
    "zachary":   "zach",
    "artyom":    "artem",
    "benjamin":  "ben",
    "joseph":    "joe",
    "matty":     "matthew",
    "aleksei":   "alexei",
    "alexey":    "alexei",
    "yegor":     "egor",
    "maxwell":   "max",
    "maxim":     "max",
}
# Reverse map: nick → formal (built from _FORMAL_TO_NICK, last formal wins on collision)
_NICK_TO_FORMAL = {v: k for k, v in _FORMAL_TO_NICK.items()}


def _add_name_aliases(d: dict, key: str, full_name: str, pg: str) -> None:
    """Emit nick/formal alias keys pointing to the same entry so either spelling matches."""
    parts = full_name.strip().split(" ", 1)
    if len(parts) < 2:
        return
    first_lower = normalize_name(parts[0]).lower()
    last = parts[1]
    for alias_first in filter(None, [_FORMAL_TO_NICK.get(first_lower), _NICK_TO_FORMAL.get(first_lower)]):
        alias_key = f"{normalize_name(alias_first + ' ' + last).lower()}_{pg}"
        if alias_key not in d:
            d[alias_key] = d[key]


@app.route("/signings")
def signings():
    import urllib.parse
    from datetime import date as _date
    scraper = _make_puckpedia_scraper()
    all_players = {}
    errors = []

    # SEASON_CUTOFF filters to only contracts still active this season.
    SEASON_CUTOFF = "2026-2027"
    # ?since=YYYY-MM-DD enables incremental mode: only fetch signings from that date onward.
    # Full sync (no since) fetches back to 2017-01-01 to catch 8-year deals signed mid-season.
    since_param = request.args.get("since")
    incremental = since_param is not None
    DATE_FROM = since_param if since_param else "2017-01-01"
    DATE_TO = "2027-12-31"

    q_base = {
        "pageSize": 100,
        "sortBy": "sign_date",
        "sortDirection": "DESC",
        "sign_date": [DATE_FROM, DATE_TO],
    }

    # First request to get total page count
    q_base["curPage"] = 1
    resp = scraper.get(
        "https://puckpedia.com/data/api_signings?q=" + urllib.parse.quote(json.dumps(q_base)),
        timeout=20,
    )
    resp.raise_for_status()
    first = resp.json()["data"]
    total_count = first["meta"]["count"]
    total_pages = (total_count + 99) // 100
    print(f"Signings: {total_count} records, {total_pages} pages")

    expired_keys = set()   # players whose most-recent contract is expired (no active deal)
    cap2526_hits = {}      # most-recent contract active during 25-26 (exp >= 2025-2026, start <= 2025)
    SEASON_2526  = "2025-2026"

    def process_page(records):
        for p in records:
            exp = p.get("exp", "")
            pos = p.get("pos", "C")
            pg = _POS_GROUP_MAP.get(pos, "F")
            full = f"{p.get('p_fn','')} {p.get('p_ln','')}".strip()
            if not full:
                continue
            first_lower = p.get("p_fn", "").lower()
            nick = _FORMAL_TO_NICK.get(first_lower)
            key = f"{normalize_name(full).lower()}_{pg}"
            nick_key = f"{normalize_name(nick + ' ' + p.get('p_ln','')).lower()}_{pg}" if nick else None

            if exp >= SEASON_CUTOFF:
                # Active contract — first occurrence wins (sign_date DESC = most recent first)
                cap = round(float(p["cap_hit"]) / 1_000_000, 4)
                if key not in all_players:
                    exp_year = int(exp.split("-")[0]) if exp else 0
                    entry = {
                        "name": full,
                        "pos":  pos,
                        "cap": cap,
                        "team": p.get("sign_team_code", ""),
                        "nhl_id": p.get("p_nhl_id", ""),
                        "term": int(p.get("len") or 0),
                        "exp_year": exp_year,
                        "birthDate": p.get("birthdate", ""),
                    }
                    all_players[key] = entry
                    if nick_key and nick_key not in all_players:
                        all_players[nick_key] = entry
            else:
                # Expired contract — only flag as expired if no active deal was found first
                if key not in all_players:
                    expired_keys.add(key)
                if nick_key and nick_key not in all_players:
                    expired_keys.add(nick_key)

            # Track 25-26 cap: most-recent contract that was active that season.
            # Sorted DESC so first match per player wins (handles multiple overlapping deals).
            try:
                term_int = int(p.get("len") or 0)
                exp_yr   = int(exp.split("-")[0]) if exp else 0
                start_yr = exp_yr - term_int + 1 if exp_yr and term_int else 0
                if exp >= SEASON_2526 and start_yr <= 2025:
                    cap_val = round(float(p["cap_hit"]) / 1_000_000, 4)
                    if key not in cap2526_hits:
                        cap2526_hits[key] = cap_val
                    if nick_key and nick_key not in cap2526_hits:
                        cap2526_hits[nick_key] = cap_val
            except (ValueError, TypeError):
                pass

    process_page(first["p"])

    for page in range(2, total_pages + 1):
        try:
            q_base["curPage"] = page
            r = scraper.get(
                "https://puckpedia.com/data/api_signings?q=" + urllib.parse.quote(json.dumps(q_base)),
                timeout=20,
            )
            r.raise_for_status()
            process_page(r.json()["data"]["p"])
            print(f"  page {page}/{total_pages} ({len(all_players)} active so far)")
            time.sleep(0.15)
        except Exception as e:
            errors.append(f"page {page}: {str(e)}")
            print(f"  ✗ page {page}: {e}")

    # Expired logic only reliable for full sync — incremental only sees recent signings
    if incremental:
        expired_keys = set()
    else:
        expired_keys -= set(all_players.keys())
    print(f"Signings {'incremental' if incremental else 'full'}: {len(all_players)} active, {len(expired_keys)} expired, {len(cap2526_hits)} 25-26 caps, {len(errors)} errors")
    return jsonify({"ok": True, "players": all_players, "expired": list(expired_keys),
                    "cap2526": cap2526_hits, "count": len(all_players), "errors": errors,
                    "fetched_through": _date.today().isoformat(),
                    "incremental": incremental})


@app.route("/signings-2526")
def signings_2526():
    """Return each player's active cap hit as of the start of the 25-26 season.

    Queries PuckPedia for contracts signed on or before 2025-09-27 (the day
    before the 25-26 season opened) whose expiry covers at least 2025-26.
    Because sign_date is sorted DESC, the first occurrence per player is their
    most-recent pre-season contract — i.e. their 25-26 cap hit.
    """
    import urllib.parse
    scraper = _make_puckpedia_scraper()
    all_players = {}
    errors = []

    SEASON_CUTOFF = "2025-2026"   # contract must cover the 25-26 season
    DATE_FROM = "2017-01-01"
    DATE_TO   = "2025-09-27"      # signed before the 25-26 season opened

    q_base = {
        "pageSize": 100,
        "sortBy": "sign_date",
        "sortDirection": "DESC",
        "sign_date": [DATE_FROM, DATE_TO],
    }

    q_base["curPage"] = 1
    resp = scraper.get(
        "https://puckpedia.com/data/api_signings?q=" + urllib.parse.quote(json.dumps(q_base)),
        timeout=20,
    )
    resp.raise_for_status()
    first = resp.json()["data"]
    total_count = first["meta"]["count"]
    total_pages = (total_count + 99) // 100
    print(f"Signings 25-26: {total_count} records, {total_pages} pages")

    def process_page(records):
        for p in records:
            exp = p.get("exp", "")
            if exp < SEASON_CUTOFF:
                continue   # contract expired before 25-26 — skip
            # Skip contracts that start in 26-27 or later (e.g. player signed new deal
            # in summer 2025 before season opened — that deal wasn't active in 25-26).
            term = int(p.get("len") or 0)
            exp_year = int(exp.split("-")[0]) if exp else 0
            start_yr = exp_year - term + 1 if term else 0
            if start_yr > 2025:
                continue
            pos = p.get("pos", "C")
            pg = _POS_GROUP_MAP.get(pos, "F")
            full = f"{p.get('p_fn','')} {p.get('p_ln','')}".strip()
            if not full:
                continue
            first_lower = p.get("p_fn", "").lower()
            nick = _FORMAL_TO_NICK.get(first_lower)
            key = f"{normalize_name(full).lower()}_{pg}"
            nick_key = f"{normalize_name(nick + ' ' + p.get('p_ln','')).lower()}_{pg}" if nick else None

            # First occurrence wins (sign_date DESC = most recent contract active in 25-26)
            cap = round(float(p["cap_hit"]) / 1_000_000, 4)
            if key not in all_players:
                entry = {"name": full, "pos": pos, "cap": cap, "nhl_id": p.get("p_nhl_id", "")}
                all_players[key] = entry
                if nick_key and nick_key not in all_players:
                    all_players[nick_key] = entry

    process_page(first["p"])

    for page in range(2, total_pages + 1):
        try:
            q_base["curPage"] = page
            r = scraper.get(
                "https://puckpedia.com/data/api_signings?q=" + urllib.parse.quote(json.dumps(q_base)),
                timeout=20,
            )
            r.raise_for_status()
            process_page(r.json()["data"]["p"])
            print(f"  page {page}/{total_pages} ({len(all_players)} active so far)")
            time.sleep(0.15)
        except Exception as e:
            errors.append(f"page {page}: {str(e)}")
            print(f"  ✗ page {page}: {e}")

    print(f"Signings 25-26 done: {len(all_players)} players, {len(errors)} errors")
    return jsonify({"ok": True, "players": all_players, "count": len(all_players), "errors": errors})


@app.route("/signings-all")
def signings_all():
    """Return all contracts from 2014 onward, grouped by player key, for client-side season assignment.

    The client determines the cap hit for each season by finding the most recently
    signed contract that was active in that season:
        cap2425 → start_yr <= 2024 AND exp >= "2024-2025"
        cap2526 → start_yr <= 2025 AND exp >= "2025-2026"
        cap     → start_yr <= 2026 AND exp >= "2026-2027"

    start_yr is computed as  exp_year − max(term, 1) + 1.
    When len = 0 (unknown term) the contract is treated as a 1-year deal so
    start_yr = exp_year — this fixes the old bug where len=0 caused start_yr=0
    and every contract passed the season filter.

    Supports incremental mode: ?since=YYYY-MM-DD fetches only signings from that
    date onward.  The client merges these into its cached store.
    """
    import urllib.parse
    from datetime import date as _date

    scraper = _make_puckpedia_scraper()

    since_param = request.args.get("since")
    DATE_FROM = since_param if since_param else "2014-01-01"
    DATE_TO   = "2028-12-31"

    q_base = {
        "pageSize": 100,
        "sortBy": "sign_date",
        "sortDirection": "DESC",
        "sign_date": [DATE_FROM, DATE_TO],
    }

    # player_key → list of contracts, ordered sign_date DESC (API order preserved)
    all_contracts: dict = {}
    errors = []
    _logged_fields = False

    def process_page(records):
        nonlocal _logged_fields
        for p in records:
            if not _logged_fields:
                print(f"  PuckPedia record fields: {sorted(p.keys())}")
                _logged_fields = True
            exp = p.get("exp", "")
            if not exp:
                continue
            try:
                exp_year = int(exp.split("-")[0])
            except (ValueError, IndexError):
                continue
            term = int(p.get("len") or 0)
            # term=0 → treat as 1-year deal so start_yr = exp_year (not 0)
            start_yr = (exp_year - term + 1) if term > 0 else exp_year

            full = f"{p.get('p_fn', '')} {p.get('p_ln', '')}".strip()
            if not full:
                continue
            cap_raw = p.get("cap_hit") or 0
            cap = round(float(cap_raw) / 1_000_000, 4)
            pos = p.get("pos", "C")
            pg  = _POS_GROUP_MAP.get(pos, "F")

            key = f"{normalize_name(full).lower()}_{pg}"
            first_lower = p.get("p_fn", "").lower()
            nick = _FORMAL_TO_NICK.get(first_lower)
            nick_key = (f"{normalize_name(nick + ' ' + p.get('p_ln', '')).lower()}_{pg}"
                        if nick else None)

            entry = {
                "name":      full,
                "pos":       pos,
                "cap":       cap,
                "sign_date": p.get("sign_date", ""),
                "exp":       exp,
                "start_yr":  start_yr,
                "term":      term,
                "nhl_id":    p.get("p_nhl_id", ""),
                "team":      p.get("sign_team_code", ""),
                "birthDate": p.get("birthdate", ""),
            }

            for k in (key, nick_key):
                if k:
                    all_contracts.setdefault(k, []).append(entry)

    q_base["curPage"] = 1
    resp = scraper.get(
        "https://puckpedia.com/data/api_signings?q=" + urllib.parse.quote(json.dumps(q_base)),
        timeout=20,
    )
    resp.raise_for_status()
    first = resp.json()["data"]
    total_count = first["meta"]["count"]
    total_pages  = (total_count + 99) // 100
    print(f"Signings all: {total_count} records, {total_pages} pages (from {DATE_FROM})")
    process_page(first["p"])

    for page in range(2, total_pages + 1):
        try:
            q_base["curPage"] = page
            r = scraper.get(
                "https://puckpedia.com/data/api_signings?q=" + urllib.parse.quote(json.dumps(q_base)),
                timeout=20,
            )
            r.raise_for_status()
            process_page(r.json()["data"]["p"])
            print(f"  page {page}/{total_pages} ({len(all_contracts)} player keys)")
            time.sleep(0.15)
        except Exception as e:
            errors.append(f"page {page}: {str(e)}")
            print(f"  ✗ page {page}: {e}")

    print(f"Signings all done: {len(all_contracts)} player keys, {len(errors)} errors")
    return jsonify({
        "ok":            True,
        "contracts":     all_contracts,
        "count":         len(all_contracts),
        "fetched_through": _date.today().isoformat(),
        "incremental":   since_param is not None,
        "errors":        errors,
    })


def _name_to_puckpedia_slug(name: str) -> str:
    """Convert a player name to a PuckPedia URL slug.

    e.g. "Ryan Nugent-Hopkins" → "ryan-nugent-hopkins"
         "Marc-André Fleury"   → "marc-andre-fleury"
         "J.T. Miller"         → "jt-miller"
    """
    import re as _re
    nfd = unicodedata.normalize("NFD", name)
    ascii_only = nfd.encode("ascii", "ignore").decode("ascii")
    slug = ascii_only.lower().replace("'", "").replace(".", "").replace(" ", "-")
    return _re.sub(r"-+", "-", slug).strip("-")


@app.route("/player-page-contracts", methods=["POST"])
def player_page_contracts():
    """Scrape PuckPedia player pages for historical contract data not in api_signings.

    api_signings only returns currently-active contracts; this endpoint fills the
    gap by parsing the Alpine.js tab_panels data embedded in each player page.

    Request body: {"players": [{"name": "Leon Draisaitl", "pos": "C"}, ...]}
    Response:     {"ok": true, "contracts": {player_key: [contract, ...]}}
    Each contract: {name, pos, cap, exp, start_yr, term, sign_date:"", nhl_id:"", team:""}
    """
    import re as _re

    req_players = request.get_json(force=True).get("players", [])
    if not req_players:
        return jsonify({"ok": True, "contracts": {}, "count": 0})

    scraper = _make_puckpedia_scraper()
    results = {}

    for item in req_players:
        name = item.get("name", "").strip()
        pos  = item.get("pos", "C")
        if not name:
            continue

        slug    = _name_to_puckpedia_slug(name)
        pg      = _POS_GROUP_MAP.get(pos, "F")
        key     = f"{normalize_name(name).lower()}_{pg}"
        first_lower = name.split()[0].lower()
        nick    = _FORMAL_TO_NICK.get(first_lower)
        last    = " ".join(name.split()[1:])
        nick_key = f"{normalize_name(nick + ' ' + last).lower()}_{pg}" if nick else None

        try:
            resp = scraper.get(f"https://puckpedia.com/player/{slug}", timeout=15)
            if resp.status_code != 200:
                print(f"  ✗ {name} ({slug}): HTTP {resp.status_code}")
                time.sleep(0.2)
                continue

            m = _re.search(
                r'id="player-contract-tab-panels".*?options:\s*(\[.*?\])\s*,\s*\n?\s*tabSelected',
                resp.text, _re.DOTALL,
            )
            if not m:
                print(f"  ✗ {name}: no contract tab panel found")
                time.sleep(0.2)
                continue

            raw_opts = m.group(1).replace("\\/", "/")
            options  = json.loads(raw_opts)
            contracts = []
            for o in options:
                title   = o.get("title", "")
                cap_str = o.get("cap_hit", "$0").replace("$", "").replace("M", "")
                try:
                    cap = round(float(cap_str), 4)
                except ValueError:
                    continue
                length = int(o.get("length") or 0)
                parts  = title.split("-")
                try:
                    start_yr = int(parts[0])
                    end_yr   = int(parts[1]) if len(parts) > 1 else start_yr + length
                except (ValueError, IndexError):
                    continue
                contracts.append({
                    "name":      name,
                    "pos":       pos,
                    "cap":       cap,
                    "sign_date": "",
                    "exp":       f"{end_yr - 1}-{end_yr}",
                    "start_yr":  start_yr,
                    "term":      length,
                    "nhl_id":    "",
                    "team":      "",
                    "birthDate": "",
                })

            if contracts:
                results[key] = contracts
                if nick_key and nick_key != key:
                    results[nick_key] = contracts
                print(f"  ✓ {name}: {len(contracts)} contracts (slug={slug})")
            else:
                print(f"  ✗ {name}: 0 contracts parsed")

            time.sleep(0.2)
        except Exception as e:
            print(f"  ✗ {name}: {e}")

    print(f"player-page-contracts done: {len(results)} player keys")
    return jsonify({"ok": True, "contracts": results, "count": len(results)})


@app.route("/injuries")
def injuries():
    """Scrape current NHL injury report from covers.com.

    Returns list of {name, key, pos, status, detail} where key is the
    slug with hyphens stripped (e.g. "troyterry") for easy JS matching.
    """
    import re as _re
    try:
        r = requests.get(
            "https://www.covers.com/sport/hockey/nhl/injuries",
            headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"},
            verify=False,
            timeout=15,
        )
        r.raise_for_status()
        html = r.text

        results = []
        pending = None

        for m in _re.finditer(r"<tr([^>]*)>(.*?)</tr>", html, _re.DOTALL):
            attrs, content = m.group(1), m.group(2)

            if 'class="collapse"' in attrs:
                if pending:
                    import html as _html
                    text = _re.sub(r"<[^>]+>", " ", content)
                    text = _re.sub(r"\s+", " ", text).strip()
                    pending["detail"] = _html.unescape(text)
                    results.append(pending)
                    pending = None
                continue

            link_m = _re.search(r"href='/sport/hockey/nhl/players/\d+/([^']+)'", content)
            if not link_m:
                pending = None
                continue

            slug = link_m.group(1)  # e.g. "troy-terry"
            key = slug.replace("-", "").lower()  # "troyterry" — matches normName strip
            display_name = " ".join(w.capitalize() for w in slug.split("-"))

            pos_m = _re.search(r"<td>\s*(C|LW|RW|W|D|G|F)\s*</td>", content)
            pos = pos_m.group(1) if pos_m else ""

            status_m = _re.search(r"<b>(.*?)</b>", content)
            status = _re.sub(r"<[^>]+>", "", status_m.group(1)).strip() if status_m else ""

            if status:
                pending = {"name": display_name, "key": key, "pos": pos, "status": status, "detail": ""}

        print(f"Injuries: {len(results)} players scraped from covers.com")
        return jsonify({"injuries": results})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  NHL Roster Tracker — Yahoo Fantasy Proxy")
    print("=" * 55)

    ensure_cert()

    tokens = load_tokens()
    if not tokens or token_expired(tokens):
        print("\n🔐  No valid token — starting OAuth flow…")
        do_oauth_flow()
    else:
        print("\n✅  Already authenticated.")

    print(f"\n🚀  Proxy running at http://localhost:{PROXY_PORT}")
    print(f"    Opening roster tracker in your browser…")
    print(f"    League: {LEAGUE_KEY}")
    print(f"    Press Ctrl+C to stop.\n")
    threading.Timer(1.5, lambda: webbrowser.open(f"http://localhost:{PROXY_PORT}")).start()
    app.run(port=PROXY_PORT, debug=False)